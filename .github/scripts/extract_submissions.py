import os
import csv
import random
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

# Load deduplicated submissions
submissions_path = "ambassador/ambassador_submissions_deduped.csv"
contrib_path = "ambassador/contribution_details.csv"

submissions = pd.read_csv(submissions_path).to_dict(orient="records")
contrib_details = pd.read_csv(contrib_path).to_dict(orient="records")

# Define reviewers
reviewers = [f"Reviewer {i}" for i in range(1, 8)]

# Updated rubric
rubric = [
    ("Technical Expertise", "Proficiency with the PyTorch Ecosystem", "Demonstrated knowledge and practical experience with PyTorch, including model building, traininga and deployment?"),
    ("Technical Expertise", "Proficiency with the PyTorch Ecosystem", "Familiarity with foundation-hosted projects, vLLM, DeepSpeed?"),
    ("Open Source Contributions", "Community Contributions", "Made commits, PRs, issues filed, and code reviews across PyTorch and its ecosystem repositories?"),
    ("Open Source Contributions", "Community Contributions", "Evidence of active participation in community discussions, RFCs, and GitHub projects?"),
    ("Open Source Contributions", "Community Contributions", "Maintenance or leadership of related open source projects or libraries?"),
    ("Thought Leadership and Technical Writing", "Publishing", "Authored technical blog posts, whitepapers, tutorials, or case studies on PyTorch or its ecosystem?"),
    ("Thought Leadership and Technical Writing", "Publishing", "Published academic research papers or publications in relevant scientific journals or conferences?"),
    ("Community Engagement and Evangelism", "Event Organization and Involvement", "Experience organizing or leading community events such as meetups, conferences, study groups, or hackathons?"),
    ("Community Engagement and Evangelism", "Event Organization and Involvement", "Participation in significant developer or ML community events (e.g., NeurIPS, PyTorch Conference, ICML, CVPR,...)"),
    ("Community Engagement and Evangelism", "Public Speaking and Presentation Skills", "Record of delivering talks, webinars, or workshops on PyTorch-related topics?"),
    ("Community Engagement and Evangelism", "Public Speaking and Presentation Skills", "Ability to communicate complex concepts clearly to both technical and non-technical audiences?"),
    ("Community Engagement and Evangelism", "Public Speaking and Presentation Skills", "Sample video recordings or links to previous talks?"),
    ("Community Engagement and Evangelism", "Mentorship and Education", "Experience mentoring students, junior developers, or researchers?"),
    ("Community Engagement and Evangelism", "Mentorship and Education", "Development or teaching of curricula or courses related to machine learning, deep learning, or distributed systems?"),
    ("Online Influence and Reach", "Social Media and Content Creation", "Active presence on platforms like Twitter, LinkedIn, YouTube, Medium, or personal blogs with a focus on machine learning, AI, or software development?"),
    ("Online Influence and Reach", "Social Media and Content Creation", "Consistency and quality of content promoting PyTorch and associated tools?"),
    ("Online Influence and Reach", "Community Impact Metrics", "High number of followers, subscribers, or consistent engagement levels with online content (>10,000 followers/>100,000 subs)?"),
    ("Online Influence and Reach", "Community Impact Metrics", "Demonstrated ability to spark discussion, share knowledge, and grow community awareness?"),
    ("Alignment and Values", "Alignment with PyTorch Foundation Values", "Commitment to open source principles, community-first development, and inclusive collaboration?"),
    ("Alignment and Values", "Alignment with PyTorch Foundation Values", "Advocacy for responsible AI development and ethical machine learning practices?"),
    ("Motivation and Vision", "Vision", "Clear articulation of why they want to be an Ambassador and what they hope to accomplish?"),
    ("Motivation and Vision", "Vision", "Proposed goals or initiatives that align with the mission of the PyTorch Foundation?"),
    ("Additional Bonus Criteria", "Cross-Community Collaboration", "Contributions or bridges to other relevant ecosystems (e.g., HuggingFace?)"),
    ("Additional Bonus Criteria", "Cross-Community Collaboration", "Integration work across tools or libraries within the AI/ML infrastructure landscape?"),
    ("Additional Bonus Criteria", "Geographic and Demographic Diversity", "Representation from underrepresented regions or groups to foster inclusivity and global outreach?"),
    ("Additional Bonus Criteria", "Innovation and Pioneering Work", "Early adoption or novel application of PyTorch or its ecosystem tools in industry, research, or startups?"),
    ("Credibility", "Community References", "References from other known community members?")
]

# Build list of unique rubric categories
summary_categories = []
for cat, _, _ in rubric:
    if cat not in summary_categories:
        summary_categories.append(cat)

# Create output folder
output_folder = "ambassador/reviewer_sheets_excel"
os.makedirs(output_folder, exist_ok=True)

# Assign reviewers randomly but evenly
assignments = []
reviewer_counts = defaultdict(int)
for submission in submissions:
    assigned = random.sample(sorted(reviewers, key=lambda r: reviewer_counts[r])[:4], 2)
    for reviewer in assigned:
        reviewer_counts[reviewer] += 1
        assignments.append((submission, reviewer))

# Generate Excel sheets
for reviewer in reviewers:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Sheet"
    summary_ws = wb.create_sheet("Score Summary")

    headers = [
        "Submission ID", "First Name", "Last Name", "Submission Summary",
        "Reviewer's Merged Comment", "Category", "Subcategory", "Question", "Score"
    ]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    dv = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    ws.add_data_validation(dv)

    row_idx = 2
    candidate_ranges = []

    for submission, assigned_reviewer in assignments:
        if assigned_reviewer != reviewer:
            continue

        sid = submission["Issue #"]
        name = submission["Nominee Name"].split()
        fname = name[0]
        lname = name[-1] if len(name) > 1 else ""

        # Pull from contribution_details
        contrib_row = next((row for row in contrib_details if int(row["Submission ID"]) == int(sid)), None)
        pitch = contrib_row.get("How Would the Nominee Contribute as an Ambassador?", "") if contrib_row else ""
        extra = contrib_row.get("Any Additional Details", "") if contrib_row else ""

        checkboxes = "\n".join([line for line in pitch.splitlines() if "☑" in line or "✔" in line])

        summary = f"""Contributions:\n{submission.get("Contributions", "").strip()}

Ambassador Pitch:\n{pitch.strip()}

Checkbox Summary:\n{checkboxes.strip()}

Extra Notes:\n{submission.get("Extra Notes", "").strip()}

Additional Info:\n{extra.strip()}"""

        start = row_idx
        for cat, subcat, question in rubric:
            ws.append([sid, fname, lname, summary, "", cat, subcat, question, ""])
            row_idx += 1
        end = row_idx - 1
        candidate_ranges.append((sid, fname, lname, start, end))

        for col in [1, 2, 3, 4, 5]:
            ws.merge_cells(start_row=start, end_row=end, start_column=col, end_column=col)
            cell = ws.cell(row=start, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        for r in range(start, end + 1):
            dv.add(ws[f"I{r}"])

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 5, 50)

    # Score Summary
    summary_ws.append(["Submission ID", "First Name", "Last Name"] + summary_categories + ["Final Score"])
    for col in range(1, summary_ws.max_column + 1):
        summary_ws.cell(row=1, column=col).font = Font(bold=True)

    for sid, fname, lname, start, end in candidate_ranges:
        category_rows = defaultdict(list)
        for r in range(start, end + 1):
            cat = ws.cell(row=r, column=6).value
            category_rows[cat].append(r)

        formulas = []
        for cat in summary_categories:
            if cat in category_rows:
                rows = category_rows[cat]
                formulas.append(f'=SUMPRODUCT(--(\'Review Sheet\'!I{rows[0]}:I{rows[-1]}="Yes"))')
            else:
                formulas.append("0")

        row_number = summary_ws.max_row + 1
        total_formula = f"=SUM({','.join([f'{get_column_letter(i+4)}{row_number}' for i in range(len(formulas))])})"
        summary_ws.append([sid, fname, lname] + formulas + [total_formula])

    filename = os.path.join(output_folder, f"{reviewer.replace(' ', '_').lower()}_sheet.xlsx")
    wb.save(filename)

print("✅ Reviewer sheets generated with enriched submission summary (including checkboxes).")
